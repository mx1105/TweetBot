#Importing modules/libraries
import tweepy
import time
import openpyxl
import tkinter as tk
from tkinter import filedialog
from time import sleep

input("Press any key and select account File.......")
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()
print(file_path)
wb_obj = openpyxl.load_workbook(file_path)
sheet_obj = wb_obj.active

# print the total number of rows
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

#Tweet URL
tweet_url = input("Enter your tweeter Post: ")
tweet_id = tweet_url.split('/')[-1]

#replies
input("Press any key and select reply File.....")
root.withdraw()
reply_file_path = filedialog.askopenfilename()
print(reply_file_path)
reply_wb_obj = openpyxl.load_workbook(reply_file_path)
reply_sheet_obj = reply_wb_obj.active
reply_max_col = reply_sheet_obj.max_column
reply_max_row = reply_sheet_obj.max_row


for i in range(1, max_row + 1, 4):
    CONSUMER_KEY = sheet_obj.cell(row = i, column = 3).value
    CONSUMER_SECRET = sheet_obj.cell(row = i+1, column = 3).value
    ACCESS_KEY = sheet_obj.cell(row = i+2, column = 3).value
    ACCESS_SECRET = sheet_obj.cell(row = i+3, column = 3).value

    auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET)
    auth.set_access_token(ACCESS_KEY,ACCESS_SECRET)
    api = tweepy.API(auth)

    og_tweet = api.get_status(tweet_id)
    print("Reply to ",og_tweet.user.screen_name)
    for j in range(1, reply_max_row + 1):
        print("Replying.......",reply_sheet_obj.cell(row = j, column = 1).value)
        api.update_status(status=reply_sheet_obj.cell(row = j, column = 1).value,in_reply_to_status_id=og_tweet.id,auto_populate_reply_metadata=True)
        sleep(0.05)
input("Press any key to exit ")
